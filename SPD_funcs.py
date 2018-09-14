import csv
import datetime
from os import path, replace, listdir, makedirs
from tkinter import filedialog
import pandas as pd

from easygui import multenterbox


from openpyxl import load_workbook


def sample_n():
    file = filedialog.askopenfilename(initialdir = "/",title = "Select your excel file",filetypes = [("Excel Files","*.xlsx")])

    msg = "Enter Arguments to create sample"
    title = "Generating random sample"
    fieldNames = ["Worksheet index: ", "Header row: ", "Skip footer: ","Number of Samples: "]
    fieldValues = []  # we start with blanks for the values
    fieldValues = multenterbox(msg, title, fieldNames)

    extract_sample(file, fieldValues[0],fieldValues[1],fieldValues[2],fieldValues[3])

def extract_sample(file, index,header,footer,n):
    if not index:
        index = 0
    if not header:
        header=0
    if not footer:
        footer =0
    if not n:
        n=25

    df = pd.read_excel(file, sheet_name=int(index)-1,header = int(header)-1,skip_rows=int(footer))

    df2 = df.sample(int(n))

    df2.reset_index(inplace=True)
    df2['index'] = df2['index'] + 1 + int(header)
    df2 = df2.rename(columns={'index': 'Old Row Number'})

    writer = pd.ExcelWriter(file, engine='openpyxl')
    book = load_workbook(file)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df2.to_excel(writer, sheet_name='New Sample', index=False)

    writer.save()

def combine_files():
    directory = filedialog.askdirectory(title="Point to directory containing files to merge:") + "/"
    counter = 0
    combined = pd.DataFrame()

    msg = "Enter parameters for File Merge"
    title = "Combine Files"
    fieldNames = ["Header Row", "Tail to skip"]
    fieldValues = []  # we start with blanks for the values
    fieldValues = multenterbox(msg, title, fieldNames)


    try:
        header_row=fieldValues[0]
        footer_rows=fieldValues[1]
    except TypeError:
        pass

    if header_row=="":
        header_row=0
    if footer_rows=="":
        footer_rows=0
    print ("header", header_row)
    print("footer", footer_rows)

    for file in listdir(directory)[:2]:
        if ".xlsx" in file:
            temp = pd.read_excel(directory + file, header=int(header_row)-1, skip_footer=int(footer_rows))
            combined = combined.append(temp, ignore_index=True)
            counter += 1
    combined.to_excel(directory+"combined_output.xlsx",index=False)
    print (counter,"files were combined and output to",directory)


def rename_all():
    directory = filedialog.askdirectory(title="Point to files to rename") + "/"

    for file in listdir(directory):
        if "0315" in file:
            temp = pd.read_excel(directory + "/" + file)
            if (len(temp.iloc[1, 0]) > 30):
                name = file[:4] + " " + temp.iloc[4, 0][15:23] + "_" + temp.iloc[5, 0][13:24] + temp.iloc[2, 0][14:16] + \
                       "-" + temp.iloc[2, 0][17:] + " to " + temp.iloc[3, 0][12:14] + "-" + temp.iloc[3, 0][15:]
            else:
                name = file[:4] + " " + temp.iloc[3, 0][15:23] + "_" + temp.iloc[4, 0][13:24] + temp.iloc[1, 0][14:16] + \
                       "-" + temp.iloc[1, 0][17:] + " to " + temp.iloc[2, 0][12:14] + "-" + temp.iloc[2, 0][15:]
            replace(directory + "/" + file, directory + "/" + name + ".xlsx")
def t_val_calc():
    df_summary = pd.DataFrame({'file': [], "total": []})
    min_dt = datetime.datetime(2018, 8, 1)
    counter = 0
    directory = filedialog.askdirectory(title="Point to a new empty directory!")+"/"
    output_directory = directory+"Tval_calc/"
    if not path.exists(output_directory):
        makedirs(output_directory)
    for file in listdir(directory):
        if (file[-5:] == '.xlsx' and "~$" not in file and "new" not in file and "summary" not in file):

            temp = pd.read_excel(directory + file)
            temp = temp[~temp.iloc[:, 2].isna()]
            temp.columns = temp.iloc[0]
            temp = temp.iloc[1:]
            # int_tot = np.sum(pd.to_numeric(temp['Interest'],errors='coerce'))
            int_tot = sum(pd.to_numeric(temp[temp['Date'] >= min_dt]['Interest'], errors='raise'))
            wb = load_workbook(directory + file)
            ws = wb.active
            ws['X5'] = "Total 08/01/2018 onwards:"
            ws['Z5'] = int_tot
            print(file, "    ", int_tot)
            wb.save(output_directory + file[:-5] + "_new.xlsx")
            df_summary.loc[counter] = [file, int_tot]
            counter += 1
    df_summary.to_excel(output_directory + "summary.xlsx", index=False)


def log_results():
    fields=[1,2,3,4,5,6,7,'a','b',8]
    with open("Resources/log.csv", 'a') as log:
        writer = csv.writer(log)
        writer.writerow(fields)
